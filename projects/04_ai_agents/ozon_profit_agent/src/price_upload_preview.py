import csv
import math
import xml.etree.ElementTree as ET
from dataclasses import dataclass
from pathlib import Path
from typing import Iterable
from zipfile import ZipFile, ZIP_DEFLATED

from openpyxl import load_workbook

from src.ozon_excel_importer import NS, _column_index, _shared_strings, _workbook_sheet_paths
from src.profit_calculator import ProfitResult
from src.report_builder import STATUS_PRIORITY

PRICE_UPLOAD_COLUMNS = {
    "offer_id": "Артикул",
    "sku": "SKU",
    "name": "Название товара",
    "old_price": "Текущая цена (со скидкой), руб.",
    "effective_price": "Цена с учетом акции или стратегии, руб.",
    "new_price": "Новая цена (со скидкой), руб.",
    "new_min_price": "Новая минимальная цена, руб.",
    "connect_promos": "Подключать подходящие акции",
    "respect_min_price": "Учитывать минимальную цену при автодобавлении в акции или продлить действие настройки",
    "auto_add_promos": "Автоматически добавлять товар в акции",
}

PRICE_PREVIEW_FIELDNAMES = [
    "status",
    "offer_id",
    "sku",
    "name",
    "current_price",
    "safe_price",
    "proposed_price",
    "price_delta",
    "profit",
    "margin_percent",
    "reason",
    "action",
]

ACTIONABLE_STATUSES = {"LOSS", "LOW_PROFIT"}


@dataclass(frozen=True)
class PriceUploadProposal:
    status: str
    offer_id: str
    sku: str
    name: str
    current_price: float
    safe_price: float
    proposed_price: int
    price_delta: float
    profit: float | None
    margin_percent: float | None
    reason: str
    action: str


def build_price_upload_proposals(results: Iterable[ProfitResult], limit: int = 5) -> list[PriceUploadProposal]:
    """Build a small, safe batch of price changes for manual review.

    Only true loss / below-target-profit items are included. Missing-cost and low-margin-only
    rows are intentionally skipped because they need human checks before price upload.
    """
    candidates = [
        item
        for item in results
        if item.status in ACTIONABLE_STATUSES and item.safe_price is not None and item.safe_price > item.current_price
    ]
    candidates.sort(
        key=lambda item: (
            STATUS_PRIORITY.get(item.status, 99),
            item.profit if item.profit is not None else 10**9,
            item.offer_id,
        )
    )

    proposals: list[PriceUploadProposal] = []
    for item in candidates[:limit]:
        proposed_price = int(math.ceil(item.safe_price))
        proposals.append(
            PriceUploadProposal(
                status=item.status,
                offer_id=item.offer_id,
                sku="",
                name=item.name,
                current_price=item.current_price,
                safe_price=item.safe_price,
                proposed_price=proposed_price,
                price_delta=round(proposed_price - item.current_price, 2),
                profit=item.profit,
                margin_percent=item.margin_percent,
                reason=item.reason,
                action="Проверить вручную; если всё верно — загрузить цену в Ozon. Акции не подключать, автодобавление в акции выключить.",
            )
        )
    return proposals


def write_price_upload_preview_csv(proposals: Iterable[PriceUploadProposal], path: str | Path) -> None:
    target = Path(path)
    target.parent.mkdir(parents=True, exist_ok=True)
    with target.open("w", encoding="utf-8", newline="") as file:
        writer = csv.DictWriter(file, fieldnames=PRICE_PREVIEW_FIELDNAMES)
        writer.writeheader()
        for item in proposals:
            writer.writerow({field: getattr(item, field) for field in PRICE_PREVIEW_FIELDNAMES})


def _header_map(sheet, header_row: int = 2) -> dict[str, int]:
    headers: dict[str, int] = {}
    for cell in sheet[header_row]:
        if cell.value:
            headers[str(cell.value).strip()] = cell.column
    return headers


def _assert_required_columns(headers: dict[str, int]) -> None:
    required = [
        PRICE_UPLOAD_COLUMNS["offer_id"],
        PRICE_UPLOAD_COLUMNS["new_price"],
        PRICE_UPLOAD_COLUMNS["new_min_price"],
        PRICE_UPLOAD_COLUMNS["connect_promos"],
        PRICE_UPLOAD_COLUMNS["respect_min_price"],
        PRICE_UPLOAD_COLUMNS["auto_add_promos"],
    ]
    missing = [column for column in required if column not in headers]
    if missing:
        raise ValueError("В Excel-файле не найдены колонки: " + ", ".join(missing))


def _column_letter(column_index: int) -> str:
    result = ""
    while column_index:
        column_index, remainder = divmod(column_index - 1, 26)
        result = chr(65 + remainder) + result
    return result


def _cell_text(cell: ET.Element, shared_strings: list[str]) -> str:
    cell_type = cell.attrib.get("t")
    if cell_type == "inlineStr":
        return "".join(text.text or "" for text in cell.findall(".//a:t", NS))
    value_node = cell.find("a:v", NS)
    if value_node is None:
        return ""
    raw_value = value_node.text or ""
    if cell_type == "s" and raw_value.isdigit():
        index = int(raw_value)
        return shared_strings[index] if index < len(shared_strings) else raw_value
    return raw_value


def _row_number(row: ET.Element) -> int:
    return int(row.attrib.get("r", "0") or 0)


def _cells_by_column(row: ET.Element) -> dict[int, ET.Element]:
    return {_column_index(cell.attrib.get("r", "A1")) + 1: cell for cell in row.findall("a:c", NS)}


def _find_or_create_cell(row: ET.Element, row_index: int, column_index: int) -> ET.Element:
    cells = _cells_by_column(row)
    if column_index in cells:
        return cells[column_index]
    cell = ET.Element(f"{{{NS['a']}}}c", {"r": f"{_column_letter(column_index)}{row_index}"})
    inserted = False
    for existing in list(row):
        if existing.tag.endswith("}c") and _column_index(existing.attrib.get("r", "A1")) + 1 > column_index:
            row.insert(list(row).index(existing), cell)
            inserted = True
            break
    if not inserted:
        row.append(cell)
    return cell


def _set_numeric_cell(cell: ET.Element, value: int | float) -> None:
    cell.attrib.pop("t", None)
    for child in list(cell):
        cell.remove(child)
    value_node = ET.SubElement(cell, f"{{{NS['a']}}}v")
    value_node.text = str(value)


def _set_inline_string_cell(cell: ET.Element, value: str) -> None:
    cell.attrib["t"] = "inlineStr"
    for child in list(cell):
        cell.remove(child)
    inline_string = ET.SubElement(cell, f"{{{NS['a']}}}is")
    text_node = ET.SubElement(inline_string, f"{{{NS['a']}}}t")
    text_node.text = value


def _write_price_upload_preview_xlsx_by_zip_edit(
    source_xlsx: str | Path,
    proposals: list[PriceUploadProposal],
    output_xlsx: str | Path,
    sheet_name: str,
) -> None:
    proposals_by_offer = {item.offer_id: item for item in proposals}
    target = Path(output_xlsx)
    target.parent.mkdir(parents=True, exist_ok=True)
    ET.register_namespace("", NS["a"])
    ET.register_namespace("r", NS["r"])

    with ZipFile(source_xlsx) as source_zip:
        shared_strings = _shared_strings(source_zip)
        sheet_paths = _workbook_sheet_paths(source_zip)
        if sheet_name not in sheet_paths:
            raise ValueError(f"В файле нет листа {sheet_name!r}")
        sheet_path = sheet_paths[sheet_name]
        root = ET.fromstring(source_zip.read(sheet_path))

        header_row = next((row for row in root.findall(".//a:sheetData/a:row", NS) if _row_number(row) == 2), None)
        if header_row is None:
            raise ValueError("В Excel-файле не найдена строка заголовков 2")
        headers = {_cell_text(cell, shared_strings).strip(): column for column, cell in _cells_by_column(header_row).items() if _cell_text(cell, shared_strings).strip()}
        _assert_required_columns(headers)

        offer_column = headers[PRICE_UPLOAD_COLUMNS["offer_id"]]
        for row in root.findall(".//a:sheetData/a:row", NS):
            row_index = _row_number(row)
            if row_index < 5:
                continue
            cells = _cells_by_column(row)
            offer_cell = cells.get(offer_column)
            offer_id = _cell_text(offer_cell, shared_strings).strip() if offer_cell is not None else ""
            proposal = proposals_by_offer.get(offer_id)
            if not proposal:
                continue
            _set_numeric_cell(_find_or_create_cell(row, row_index, headers[PRICE_UPLOAD_COLUMNS["new_price"]]), proposal.proposed_price)
            _set_numeric_cell(_find_or_create_cell(row, row_index, headers[PRICE_UPLOAD_COLUMNS["new_min_price"]]), proposal.proposed_price)
            _set_inline_string_cell(_find_or_create_cell(row, row_index, headers[PRICE_UPLOAD_COLUMNS["connect_promos"]]), "Нет")
            _set_inline_string_cell(_find_or_create_cell(row, row_index, headers[PRICE_UPLOAD_COLUMNS["respect_min_price"]]), "Да")
            _set_inline_string_cell(_find_or_create_cell(row, row_index, headers[PRICE_UPLOAD_COLUMNS["auto_add_promos"]]), "Нет")

        modified_sheet = ET.tostring(root, encoding="utf-8", xml_declaration=True)
        with ZipFile(target, "w", ZIP_DEFLATED) as output_zip:
            for item in source_zip.infolist():
                output_zip.writestr(item, modified_sheet if item.filename == sheet_path else source_zip.read(item.filename))


def write_price_upload_preview_xlsx(
    source_xlsx: str | Path,
    proposals: Iterable[PriceUploadProposal],
    output_xlsx: str | Path,
    sheet_name: str = "Товары и цены",
) -> None:
    proposals = list(proposals)
    try:
        _write_price_upload_preview_xlsx_with_openpyxl(source_xlsx, proposals, output_xlsx, sheet_name)
    except ValueError as error:
        if "stylesheet" not in str(error).lower() and "invalid xml" not in str(error).lower():
            raise
        _write_price_upload_preview_xlsx_by_zip_edit(source_xlsx, proposals, output_xlsx, sheet_name)


def _write_price_upload_preview_xlsx_with_openpyxl(
    source_xlsx: str | Path,
    proposals: list[PriceUploadProposal],
    output_xlsx: str | Path,
    sheet_name: str,
) -> None:
    proposals_by_offer = {item.offer_id: item for item in proposals}
    target = Path(output_xlsx)
    target.parent.mkdir(parents=True, exist_ok=True)

    workbook = load_workbook(source_xlsx)
    if sheet_name not in workbook.sheetnames:
        raise ValueError(f"В файле нет листа {sheet_name!r}")
    sheet = workbook[sheet_name]
    headers = _header_map(sheet)

    _assert_required_columns(headers)

    for row_index in range(5, sheet.max_row + 1):
        offer_id = str(sheet.cell(row=row_index, column=headers[PRICE_UPLOAD_COLUMNS["offer_id"]]).value or "").strip()
        proposal = proposals_by_offer.get(offer_id)
        if not proposal:
            continue
        if PRICE_UPLOAD_COLUMNS["sku"] in headers:
            proposal = PriceUploadProposal(
                status=proposal.status,
                offer_id=proposal.offer_id,
                sku=str(sheet.cell(row=row_index, column=headers[PRICE_UPLOAD_COLUMNS["sku"]]).value or ""),
                name=proposal.name,
                current_price=proposal.current_price,
                safe_price=proposal.safe_price,
                proposed_price=proposal.proposed_price,
                price_delta=proposal.price_delta,
                profit=proposal.profit,
                margin_percent=proposal.margin_percent,
                reason=proposal.reason,
                action=proposal.action,
            )
            proposals_by_offer[offer_id] = proposal
        sheet.cell(row=row_index, column=headers[PRICE_UPLOAD_COLUMNS["new_price"]]).value = proposal.proposed_price
        sheet.cell(row=row_index, column=headers[PRICE_UPLOAD_COLUMNS["new_min_price"]]).value = proposal.proposed_price
        sheet.cell(row=row_index, column=headers[PRICE_UPLOAD_COLUMNS["connect_promos"]]).value = "Нет"
        sheet.cell(row=row_index, column=headers[PRICE_UPLOAD_COLUMNS["respect_min_price"]]).value = "Да"
        sheet.cell(row=row_index, column=headers[PRICE_UPLOAD_COLUMNS["auto_add_promos"]]).value = "Нет"

    workbook.save(target)
