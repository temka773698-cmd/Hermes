import csv

from openpyxl import Workbook, load_workbook

from src.price_upload_preview import (
    build_price_upload_proposals,
    write_price_upload_preview_csv,
    write_price_upload_preview_xlsx,
)
from src.profit_calculator import ProfitResult


def test_build_price_upload_proposals_limits_to_top_loss_and_low_profit():
    results = [
        ProfitResult("OK-1", "Нормальный", 1000, 900, 200, 20, "OK", "", ""),
        ProfitResult("LOW-MARGIN", "Низкая маржа", 1000, 900, 100, 10, "LOW_MARGIN", "", ""),
        ProfitResult("LOW-1", "Мало прибыли", 900, 1100.2, 100, 11, "LOW_PROFIT", "", ""),
        ProfitResult("LOSS-1", "Минус", 1000, 1200.1, -100, -10, "LOSS", "", ""),
        ProfitResult("LOSS-2", "Минус сильнее", 1000, 1300.1, -200, -20, "LOSS", "", ""),
    ]

    proposals = build_price_upload_proposals(results, limit=2)

    assert [item.offer_id for item in proposals] == ["LOSS-2", "LOSS-1"]
    assert proposals[0].proposed_price == 1301
    assert proposals[0].price_delta == 301


def test_write_price_upload_preview_csv(tmp_path):
    proposal = build_price_upload_proposals(
        [ProfitResult("LOSS-1", "Минус", 1000, 1200.1, -100, -10, "LOSS", "Причина", "Рекомендация")]
    )[0]
    path = tmp_path / "preview.csv"

    write_price_upload_preview_csv([proposal], path)

    rows = list(csv.DictReader(path.open(encoding="utf-8")))
    assert rows[0]["offer_id"] == "LOSS-1"
    assert rows[0]["proposed_price"] == "1201"
    assert rows[0]["status"] == "LOSS"


def test_write_price_upload_preview_xlsx_fills_only_selected_rows(tmp_path):
    source = tmp_path / "source.xlsx"
    output = tmp_path / "output.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Товары и цены"
    sheet.append(["служебная строка"])
    sheet.append([
        "Артикул",
        "SKU",
        "Название товара",
        "Текущая цена (со скидкой), руб.",
        "Новая цена (со скидкой), руб.",
        "Новая минимальная цена, руб.",
        "Подключать подходящие акции",
        "Учитывать минимальную цену при автодобавлении в акции или продлить действие настройки",
        "Автоматически добавлять товар в акции",
    ])
    sheet.append([])
    sheet.append([])
    sheet.append(["LOSS-1", "123", "Минус", 1000, None, None, None, None, None])
    sheet.append(["OK-1", "456", "Норма", 1000, None, None, None, None, None])
    workbook.save(source)
    proposal = build_price_upload_proposals(
        [ProfitResult("LOSS-1", "Минус", 1000, 1200.1, -100, -10, "LOSS", "Причина", "Рекомендация")]
    )[0]

    write_price_upload_preview_xlsx(source, [proposal], output)

    result = load_workbook(output)
    sheet = result["Товары и цены"]
    assert sheet.cell(row=5, column=5).value == 1201
    assert sheet.cell(row=5, column=6).value == 1201
    assert sheet.cell(row=5, column=7).value == "Нет"
    assert sheet.cell(row=5, column=8).value == "Да"
    assert sheet.cell(row=5, column=9).value == "Нет"
    assert sheet.cell(row=6, column=5).value is None
